import datetime
import pandas as pd
import numpy as np
import akshare as ak
import matplotlib.pyplot as plt
import backtrader as bt
import os
import time
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 设置中文显示
plt.rcParams["font.sans-serif"] = ["SimHei"]
plt.rcParams["axes.unicode_minus"] = False

class BuySignalAnalyzer:
    """
    买入信号分析器：分析股票是否触发买入信号
    基于HighReturnStrategy的买入策略
    """
    def __init__(self, stock_code, stock_name=None, data_days=120, include_today=True):
        """
        初始化分析器
        
        参数:
        stock_code: 股票代码
        stock_name: 股票名称（可选）
        data_days: 获取的历史数据天数
        include_today: 是否包含今天的数据
        """
        self.stock_code = stock_code
        self.stock_name = stock_name if stock_name else stock_code
        self.data_days = data_days
        self.include_today = include_today
        self.data = None
        self.buy_signals = []
        self.score = 0  # 综合得分
        
        # 策略参数
        self.params = {
            "sma5": 5,
            "sma10": 10,
            "sma20": 20,
            "sma60": 60,
            "macd_fast": 12,
            "macd_slow": 26,
            "macd_signal": 9,
            "rsi_period": 14,
            "rsi_oversold": 30,
            "volume_ratio": 1.5,
            "ma_convergence_threshold": 0.001,  # 均线聚合阈值
            "ma_rising_threshold": 0.002,      # 均线上涨阈值
            # 新增指标参数
            "kdj_period": 9,
            "kdj_signal_period": 3,
            "bias_period": 26,
            "cci_period": 14,
            "atr_period": 14,
            "obv_period": 20,
        }
        
        # 评分权重
        self.weights = {
            "强势突破": 30,
            "低吸回调": 25,
            "三线合一上涨": 20,  # 修改为三线合一上涨
            "KDJ金叉": 5,
            "BIAS回归": 5,
            "CCI超买超卖": 5,
            "OBV增长": 5,
            "量价配合": 5
        }
        
    def load_data(self):
        """加载股票数据"""
        try:
            print(f"尝试加载股票 {self.stock_code} 的数据...")
            # 使用AKShare获取股票数据
            stock_df = ak.stock_zh_a_hist(symbol=self.stock_code, adjust="qfq", period="daily")
            
            if stock_df is None or len(stock_df) == 0:
                print(f"股票 {self.stock_code} 数据为空")
                return False
                
            print(f"成功获取股票 {self.stock_code} 的数据, 共 {len(stock_df)} 行")
            
            # 只保留需要的列
            stock_df = stock_df[['日期', '开盘', '收盘', '最高', '最低', '成交量']]
            stock_df.columns = ['date', 'open', 'close', 'high', 'low', 'volume']
            
            # 转换日期列为日期类型
            stock_df['date'] = pd.to_datetime(stock_df['date'])
            
            # 如果不包含今天的数据，则排除最后一行
            if not self.include_today:
                today = pd.Timestamp.now().normalize()
                # 检查是否有今天的数据
                has_today = (stock_df['date'].max() == today)
                if has_today:
                    print(f"根据设置，排除今天({today.strftime('%Y-%m-%d')})的数据")
                    stock_df = stock_df[stock_df['date'] < today]
            
            # 只保留最近N天的数据
            if len(stock_df) > self.data_days:
                stock_df = stock_df.tail(self.data_days)
                
            # 设置日期为索引
            stock_df.set_index('date', inplace=True)
            
            self.data = stock_df
            return True
        except Exception as e:
            print(f"加载股票 {self.stock_code} 数据失败: {e}")
            # 尝试显示更多的错误信息
            import traceback
            traceback.print_exc()
            return False
    
    def calculate_indicators(self):
        """计算技术指标"""
        if self.data is None or len(self.data) < 60:
            print(f"股票 {self.stock_code} 数据不足，无法计算指标")
            return False
            
        # 计算移动平均线
        self.data['sma5'] = self.data['close'].rolling(window=self.params['sma5']).mean()
        self.data['sma10'] = self.data['close'].rolling(window=self.params['sma10']).mean()
        self.data['sma20'] = self.data['close'].rolling(window=self.params['sma20']).mean()
        self.data['sma60'] = self.data['close'].rolling(window=self.params['sma60']).mean()
        
        # 计算均线斜率
        self.data['sma5_slope'] = self.data['sma5'].pct_change(1)
        self.data['sma10_slope'] = self.data['sma10'].pct_change(1)
        self.data['sma20_slope'] = self.data['sma20'].pct_change(1)
        
        # 计算MACD
        exp1 = self.data['close'].ewm(span=self.params['macd_fast'], adjust=False).mean()
        exp2 = self.data['close'].ewm(span=self.params['macd_slow'], adjust=False).mean()
        self.data['macd'] = exp1 - exp2
        self.data['macd_signal'] = self.data['macd'].ewm(span=self.params['macd_signal'], adjust=False).mean()
        self.data['macd_hist'] = self.data['macd'] - self.data['macd_signal']
        
        # 计算RSI
        delta = self.data['close'].diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=self.params['rsi_period']).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=self.params['rsi_period']).mean()
        rs = gain / loss
        self.data['rsi'] = 100 - (100 / (1 + rs))
        
        # 计算布林带
        self.data['bb_middle'] = self.data['close'].rolling(window=20).mean()
        std = self.data['close'].rolling(window=20).std()
        self.data['bb_upper'] = self.data['bb_middle'] + 2 * std
        self.data['bb_lower'] = self.data['bb_middle'] - 2 * std
        
        # 计算成交量均线
        self.data['volume_ma20'] = self.data['volume'].rolling(window=20).mean()
        
        # =============== 新增技术指标计算 ===============
        
        # 1. 计算KDJ指标
        low_min = self.data['low'].rolling(window=self.params['kdj_period']).min()
        high_max = self.data['high'].rolling(window=self.params['kdj_period']).max()
        
        # 计算RSV
        rsv = 100 * ((self.data['close'] - low_min) / (high_max - low_min + 1e-10))
        
        # 计算K、D、J值
        self.data['kdj_k'] = rsv.ewm(alpha=1/3, adjust=False).mean()
        self.data['kdj_d'] = self.data['kdj_k'].ewm(alpha=1/3, adjust=False).mean()
        self.data['kdj_j'] = 3 * self.data['kdj_k'] - 2 * self.data['kdj_d']
        
        # 2. 计算BIAS乖离率
        self.data['bias'] = 100 * (self.data['close'] - self.data['close'].rolling(window=self.params['bias_period']).mean()) / self.data['close'].rolling(window=self.params['bias_period']).mean()
        
        # 3. 计算OBV能量潮
        self.data['obv'] = 0
        self.data.iloc[0, self.data.columns.get_loc('obv')] = self.data.iloc[0, self.data.columns.get_loc('volume')]
        for i in range(1, len(self.data)):
            if self.data.iloc[i]['close'] > self.data.iloc[i-1]['close']:
                self.data.iloc[i, self.data.columns.get_loc('obv')] = self.data.iloc[i-1, self.data.columns.get_loc('obv')] + self.data.iloc[i, self.data.columns.get_loc('volume')]
            elif self.data.iloc[i]['close'] < self.data.iloc[i-1]['close']:
                self.data.iloc[i, self.data.columns.get_loc('obv')] = self.data.iloc[i-1, self.data.columns.get_loc('obv')] - self.data.iloc[i, self.data.columns.get_loc('volume')]
            else:
                self.data.iloc[i, self.data.columns.get_loc('obv')] = self.data.iloc[i-1, self.data.columns.get_loc('obv')]
        
        # OBV移动平均线
        self.data['obv_ma'] = self.data['obv'].rolling(window=self.params['obv_period']).mean()
        
        # 4. 计算CCI指标
        tp = (self.data['high'] + self.data['low'] + self.data['close']) / 3  # 典型价格
        tp_ma = tp.rolling(window=self.params['cci_period']).mean()  # 典型价格移动平均
        md = (tp - tp_ma).abs().rolling(window=self.params['cci_period']).mean()  # 平均偏差
        self.data['cci'] = (tp - tp_ma) / (0.015 * md)
        
        # 5. 计算ATR平均真实波幅
        tr1 = self.data['high'] - self.data['low']
        tr2 = (self.data['high'] - self.data['close'].shift(1)).abs()
        tr3 = (self.data['low'] - self.data['close'].shift(1)).abs()
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
        self.data['atr'] = tr.rolling(window=self.params['atr_period']).mean()
        
        # 量价关系指标：成交量变化率
        self.data['volume_change'] = self.data['volume'].pct_change(1)
        self.data['price_volume_ratio'] = self.data['close'].pct_change(1) / (self.data['volume_change'] + 1e-10)
        
        return True
    
    def analyze_buy_signals(self):
        """分析买入信号"""
        if self.data is None or 'sma5' not in self.data.columns:
            print(f"股票 {self.stock_code} 指标未计算，无法分析买入信号")
            return False
            
        # 获取最新的数据点
        latest = self.data.iloc[-1]
        prev = self.data.iloc[-2] if len(self.data) > 1 else None
        
        # 初始化信号字典
        signals = {
            "强势突破": False,
            "低吸回调": False,
            "三线合一上涨": False,
            "信号详情": {}
        }
        
        # 1. 强势突破买入信号
        golden_cross = latest['sma5'] > latest['sma10'] > latest['sma20']
        all_ma_rising = (latest['sma5_slope'] > 0 and latest['sma10_slope'] > 0 and latest['sma20_slope'] > 0)
        price_above_all_ma = (latest['close'] > latest['sma5'] and latest['close'] > latest['sma10'])
        
        # MACD金叉
        macd_golden_cross = False
        if prev is not None:
            macd_golden_cross = (latest['macd'] > latest['macd_signal'] and 
                                prev['macd'] <= prev['macd_signal'])
        
        # 成交量确认
        volume_confirm = latest['volume'] > latest['volume_ma20'] * self.params['volume_ratio']
        
        # 强势突破信号
        breakout_signal = (golden_cross and 
                          all_ma_rising and 
                          price_above_all_ma and 
                          (macd_golden_cross or volume_confirm))
        
        signals["强势突破"] = breakout_signal
        signals["信号详情"]["强势突破"] = {
            "三线黄金排列": golden_cross,
            "均线同向上涨": all_ma_rising,
            "价格站上均线": price_above_all_ma,
            "MACD金叉": macd_golden_cross,
            "成交量确认": volume_confirm
        }
        
        # 2. 低吸回调买入信号
        rsi_oversold = latest['rsi'] < self.params['rsi_oversold']
        near_lower_band = latest['close'] < latest['bb_lower'] * 1.05
        
        # 价格企稳
        price_stabilizing = False
        if prev is not None:
            price_stabilizing = (latest['close'] > latest['open'] and 
                                prev['close'] > prev['open'])
        
        # 价格反弹确认
        price_bounce = False
        if len(self.data) > 2:
            prev2 = self.data.iloc[-3]
            price_bounce = (latest['close'] > latest['open'] and 
                           prev['close'] > prev['open'] and
                           latest['close'] > prev['close'])
        
        # 判断趋势
        trend = 'unknown'
        if latest['sma5'] > latest['sma10'] > latest['sma20']:
            if latest['sma5_slope'] > 0 and latest['sma10_slope'] > 0:
                trend = 'up'
            else:
                trend = 'sideways'
        elif latest['sma5'] < latest['sma10'] < latest['sma20']:
            if latest['sma5_slope'] < 0 and latest['sma10_slope'] < 0:
                trend = 'down'
            else:
                trend = 'sideways'
        else:
            trend = 'sideways'
        
        # 低吸回调信号
        dip_buy_signal = (
            (rsi_oversold or near_lower_band) and 
            (price_stabilizing or price_bounce) and
            trend != 'down'  # 不在下跌趋势中
        )
        
        signals["低吸回调"] = dip_buy_signal
        signals["信号详情"]["低吸回调"] = {
            "RSI超卖": rsi_oversold,
            "接近布林带下轨": near_lower_band,
            "价格企稳": price_stabilizing,
            "价格反弹": price_bounce,
            "非下跌趋势": trend != 'down'
        }
        
        # 3. 三线合一上涨买入信号
        # 计算均线之间的距离
        ma5_ma10_distance = abs(latest['sma5'] - latest['sma10']) / latest['sma5']
        ma10_ma20_distance = abs(latest['sma10'] - latest['sma20']) / latest['sma10']
        
        # 均线聚合条件
        ma_convergence = (ma5_ma10_distance < self.params['ma_convergence_threshold'] and 
                         ma10_ma20_distance < self.params['ma_convergence_threshold'])
        
        # 均线方向向上
        ma5_rising = latest['sma5_slope'] > self.params['ma_rising_threshold']
        ma10_rising = latest['sma10_slope'] > self.params['ma_rising_threshold']
        ma20_rising = latest['sma20_slope'] > self.params['ma_rising_threshold']
        
        # 价格站上所有均线
        price_above_mas = (latest['close'] > latest['sma5'] and 
                          latest['close'] > latest['sma10'] and 
                          latest['close'] > latest['sma20'])
        
        # 三线合一上涨信号
        three_line_up_signal = (
            ma_convergence and  # 均线聚合
            ma5_rising and      # 5日均线向上
            ma10_rising and     # 10日均线向上
            ma20_rising and     # 20日均线向上
            price_above_mas     # 价格站上均线
        )
        
        signals["三线合一上涨"] = three_line_up_signal
        signals["信号详情"]["三线合一上涨"] = {
            "均线聚合": ma_convergence,
            "MA5-MA10距离": ma5_ma10_distance,
            "MA10-MA20距离": ma10_ma20_distance,
            "5日均线向上": ma5_rising,
            "10日均线向上": ma10_rising,
            "20日均线向上": ma20_rising,
            "价格站上均线": price_above_mas
        }
        
        # 分析新增的技术指标信号
        self.analyze_additional_signals(signals, latest, prev)
        
        # 综合买入信号
        signals["任一买入信号"] = signals["强势突破"] or signals["低吸回调"] or signals["三线合一上涨"] or signals["KDJ金叉"] or signals["BIAS回归"] or signals["CCI超买超卖"] or signals["OBV增长"] or signals["量价配合"]
        
        # 计算综合得分
        self.calculate_score(signals)
        
        self.buy_signals = signals
        return True
    
    def analyze_additional_signals(self, signals, latest, prev):
        """分析新增的技术指标信号"""
        if prev is None:
            return

        # 初始化新信号
        signals["KDJ金叉"] = False
        signals["BIAS回归"] = False
        signals["CCI超买超卖"] = False
        signals["OBV增长"] = False
        signals["量价配合"] = False
        signals["信号详情"]["KDJ金叉"] = {}
        signals["信号详情"]["BIAS回归"] = {}
        signals["信号详情"]["CCI超买超卖"] = {}
        signals["信号详情"]["OBV增长"] = {}
        signals["信号详情"]["量价配合"] = {}
        
        # 1. KDJ金叉信号
        kdj_golden_cross = (latest['kdj_j'] > latest['kdj_d'] and 
                           prev['kdj_j'] <= prev['kdj_d'] and
                           latest['kdj_j'] < 80)  # 避免超买区间的金叉
        
        kdj_oversold_bounce = (latest['kdj_j'] < 20 and latest['kdj_j'] > prev['kdj_j'])
        
        signals["KDJ金叉"] = kdj_golden_cross or kdj_oversold_bounce
        signals["信号详情"]["KDJ金叉"] = {
            "KDJ金叉": kdj_golden_cross,
            "KDJ超卖反弹": kdj_oversold_bounce,
            "J值": latest['kdj_j'],
            "D值": latest['kdj_d']
        }
        
        # 2. BIAS乖离率信号
        bias_return_to_zero = abs(latest['bias']) < 3.0 and abs(prev['bias']) >= 3.0
        bias_oversold_bounce = latest['bias'] < -6.0 and latest['bias'] > prev['bias']
        
        signals["BIAS回归"] = bias_return_to_zero or bias_oversold_bounce
        signals["信号详情"]["BIAS回归"] = {
            "BIAS回归0轴": bias_return_to_zero,
            "BIAS超卖反弹": bias_oversold_bounce,
            "BIAS值": latest['bias']
        }
        
        # 3. CCI指标信号
        cci_oversold_bounce = latest['cci'] < -100 and latest['cci'] > prev['cci']
        cci_return_to_zero = abs(latest['cci']) < 50 and abs(prev['cci']) >= 100
        
        signals["CCI超买超卖"] = cci_oversold_bounce or cci_return_to_zero
        signals["信号详情"]["CCI超买超卖"] = {
            "CCI超卖反弹": cci_oversold_bounce,
            "CCI回归0轴": cci_return_to_zero,
            "CCI值": latest['cci']
        }
        
        # 4. OBV能量潮信号
        obv_rising = latest['obv'] > latest['obv_ma'] and latest['obv'] > prev['obv']
        obv_rising_price_stable = obv_rising and latest['close'] >= prev['close']
        
        signals["OBV增长"] = obv_rising_price_stable
        signals["信号详情"]["OBV增长"] = {
            "OBV上升": obv_rising,
            "OBV上升价格稳定": obv_rising_price_stable,
            "OBV值": latest['obv'],
            "OBV均线值": latest['obv_ma']
        }
        
        # 5. 量价配合信号
        volume_price_positive = latest['price_volume_ratio'] > 0 and latest['close'] > prev['close'] and latest['volume'] > prev['volume']
        volume_increasing = latest['volume'] > latest['volume_ma20'] * 1.2
        
        signals["量价配合"] = volume_price_positive and volume_increasing
        signals["信号详情"]["量价配合"] = {
            "量增价涨": volume_price_positive,
            "成交量放大": volume_increasing,
            "量价比": latest['price_volume_ratio'],
            "成交量变化率": latest['volume_change']
        }
        
    def calculate_score(self, signals):
        """计算综合评分"""
        score = 0
        
        # 基于信号计算得分
        for signal_name, weight in self.weights.items():
            if signal_name in signals and signals[signal_name]:
                score += weight
                
        # 获取最新数据
        latest = self.data.iloc[-1]
        
        # 加分项：根据交易量、价格和指标走势增加额外得分
        
        # 1. MACD方向加分
        if 'macd_hist' in self.data.columns and len(self.data) > 1:
            prev = self.data.iloc[-2]
            if latest['macd_hist'] > 0 and latest['macd_hist'] > prev['macd_hist']:
                score += 3
                
        # 2. RSI状态加分
        if 'rsi' in self.data.columns:
            if 30 <= latest['rsi'] <= 70:  # 健康区间
                score += 2
            elif latest['rsi'] > 70:  # 超买区间扣分
                score -= 2
                
        # 3. 均线方向加分
        if all(latest[f'sma{period}_slope'] > 0 for period in [5, 10, 20]):
            score += 3
            
        # 4. ATR波动率考虑
        if 'atr' in self.data.columns and 'close' in self.data.columns:
            atr_percent = latest['atr'] / latest['close'] * 100
            if atr_percent < 2:  # 波动率较低
                score += 2
                
        # 5. 价格位置考虑
        if 'bb_middle' in self.data.columns and 'bb_upper' in self.data.columns:
            price_position = (latest['close'] - latest['bb_lower']) / (latest['bb_upper'] - latest['bb_lower'])
            if 0.3 <= price_position <= 0.7:  # 价格在布林带中间位置
                score += 2
            elif price_position > 0.8:  # 接近上轨扣分
                score -= 2
                
        # 保存得分
        self.score = score
    
    def get_result(self):
        """获取分析结果"""
        if not self.buy_signals:
            return {
                "股票代码": self.stock_code,
                "股票名称": self.stock_name,
                "任一买入信号": False,
                "强势突破": False,
                "低吸回调": False,
                "三线合一上涨": False,
                "KDJ金叉": False,
                "BIAS回归": False,
                "CCI超买超卖": False,
                "OBV增长": False,
                "量价配合": False,
                "综合评分": 0,
                "最新价格": None,
                "信号详情": {}
            }
        
        latest_price = self.data['close'].iloc[-1] if self.data is not None and len(self.data) > 0 else None
        
        # 提取最新技术指标值
        latest_indicators = {}
        if self.data is not None and len(self.data) > 0:
            latest = self.data.iloc[-1]
            latest_indicators = {
                "RSI": latest.get('rsi', None),
                "MACD": latest.get('macd', None),
                "KDJ_J": latest.get('kdj_j', None),
                "BIAS": latest.get('bias', None),
                "CCI": latest.get('cci', None),
                "ATR": latest.get('atr', None),
                "OBV": latest.get('obv', None),
                "SMA5": latest.get('sma5', None),
                "SMA10": latest.get('sma10', None),
                "SMA20": latest.get('sma20', None),
                "SMA60": latest.get('sma60', None)
            }
        
        return {
            "股票代码": self.stock_code,
            "股票名称": self.stock_name,
            "任一买入信号": self.buy_signals["任一买入信号"],
            "强势突破": self.buy_signals["强势突破"],
            "低吸回调": self.buy_signals["低吸回调"],
            "三线合一上涨": self.buy_signals["三线合一上涨"],
            "KDJ金叉": self.buy_signals.get("KDJ金叉", False),
            "BIAS回归": self.buy_signals.get("BIAS回归", False),
            "CCI超买超卖": self.buy_signals.get("CCI超买超卖", False),
            "OBV增长": self.buy_signals.get("OBV增长", False),
            "量价配合": self.buy_signals.get("量价配合", False),
            "综合评分": self.score,
            "最新价格": latest_price,
            "信号详情": self.buy_signals["信号详情"],
            "技术指标": latest_indicators
        }
    
    def run_analysis(self):
        """运行完整分析流程"""
        if not self.load_data():
            return self.get_result()
        
        if not self.calculate_indicators():
            return self.get_result()
        
        self.analyze_buy_signals()
        return self.get_result()


class StockScanner:
    """
    股票扫描器：扫描股票列表，分析哪些股票触发了买入信号
    """
    def __init__(self, stock_list=None, use_index=False, index_code="000300", top_n=300, new_account_only=False, include_today=True):
        """
        初始化扫描器
        
        参数:
        stock_list: 股票列表，格式为 [(code, name), ...]
        use_index: 是否使用指数成分股
        index_code: 指数代码，默认为沪深300
        top_n: 使用指数成分股时，选取的股票数量
        new_account_only: 是否只筛选新开户投资者可买的股票（00、000、002、60开头）
        include_today: 是否包含今天的数据
        """
        self.stock_list = stock_list if stock_list else []
        self.use_index = use_index
        self.index_code = index_code
        self.top_n = top_n
        self.new_account_only = new_account_only
        self.include_today = include_today
        self.results = []
        
    def is_new_account_eligible(self, stock_code):
        """
        判断股票是否符合新开户投资者可买条件
        
        参数:
        stock_code: 股票代码
        
        返回:
        bool: 是否符合条件
        """
        if not stock_code:
            return False
            
        # 去掉股票代码中可能的前缀（如sh.、sz.等）
        code = stock_code.strip()
        if '.' in code:
            code = code.split('.')[-1]
            
        # 判断是否为新开户可买股票（00、000、002、60开头）
        return (code.startswith('00') or 
                code.startswith('000') or 
                code.startswith('002') or 
                code.startswith('60'))
    
    def load_new_account_stocks(self, limit=None):
        """
        直接加载所有新开户可买的股票
        
        参数:
        limit: 限制加载的股票数量
        
        返回:
        bool: 是否成功加载
        """
        try:
            print("正在获取所有A股股票列表...")
            # 获取A股股票列表
            stock_list = ak.stock_zh_a_spot_em()
            # 只保留股票代码和名称
            stock_list = stock_list[['代码', '名称']].values.tolist()
            
            # 筛选出新开户可买的股票（00、000、002、60开头）
            new_account_stocks = [stock for stock in stock_list 
                                  if self.is_new_account_eligible(stock[0])]
            
            print(f"共找到 {len(new_account_stocks)} 只新开户可买股票")
            
            # 限制数量
            if limit and len(new_account_stocks) > limit:
                new_account_stocks = new_account_stocks[:limit]
                print(f"已限制为前 {limit} 只股票")
            
            self.stock_list = new_account_stocks
            return True
        except Exception as e:
            print(f"获取新开户可买股票失败: {e}")
            import traceback
            traceback.print_exc()
            return False
        
    def load_stock_list(self):
        """加载股票列表"""
        if self.stock_list:
            # 如果启用新开户筛选，筛选符合条件的股票
            if self.new_account_only:
                self.stock_list = [stock for stock in self.stock_list 
                                   if self.is_new_account_eligible(stock[0])]
            return
            
        if self.use_index:
            try:
                # 获取指数成分股
                if self.index_code == "000300":
                    # 沪深300成分股
                    index_stocks = ak.index_stock_cons_weight_csindex(symbol="000300")
                    # 只保留股票代码和名称
                    index_stocks = index_stocks[['成分券代码', '成分券名称']].values.tolist()
                    
                    # 如果启用新开户筛选，筛选符合条件的股票
                    if self.new_account_only:
                        index_stocks = [stock for stock in index_stocks 
                                        if self.is_new_account_eligible(stock[0])]
                    
                    # 只保留前N个
                    self.stock_list = index_stocks[:self.top_n]
                else:
                    # 其他指数可以根据需要添加
                    print(f"暂不支持指数 {self.index_code} 的成分股获取")
                    self.stock_list = []
            except Exception as e:
                print(f"获取指数成分股失败: {e}")
                self.stock_list = []
        else:
            try:
                # 获取A股股票列表
                stock_list = ak.stock_zh_a_spot_em()
                # 只保留股票代码和名称
                stock_list = stock_list[['代码', '名称']].values.tolist()
                
                # 如果启用新开户筛选，筛选符合条件的股票
                if self.new_account_only:
                    stock_list = [stock for stock in stock_list 
                                  if self.is_new_account_eligible(stock[0])]
                
                # 只保留前N个
                self.stock_list = stock_list[:self.top_n]
            except Exception as e:
                print(f"获取A股股票列表失败: {e}")
                self.stock_list = []
    
    def analyze_single_stock(self, stock):
        """分析单个股票"""
        try:
            code, name = stock
            analyzer = BuySignalAnalyzer(code, name, include_today=self.include_today)
            result = analyzer.run_analysis()
            return result
        except Exception as e:
            print(f"处理股票 {stock} 时出错: {e}")
            # 返回一个带有基本信息但无买入信号的结果
            if isinstance(stock, (list, tuple)) and len(stock) >= 2:
                code, name = stock[0], stock[1]
            else:
                code, name = str(stock), str(stock)
            return {
                "股票代码": code,
                "股票名称": name,
                "任一买入信号": False,
                "强势突破": False,
                "低吸回调": False,
                "三线合一上涨": False,
                "KDJ金叉": False,
                "BIAS回归": False,
                "CCI超买超卖": False,
                "OBV增长": False,
                "量价配合": False,
                "综合评分": 0,
                "最新价格": None,
                "信号详情": {}
            }
    
    def scan_stocks(self, max_workers=10):
        """扫描所有股票"""
        self.load_stock_list()
        
        if not self.stock_list:
            print("股票列表为空，无法进行扫描")
            return []
        
        print(f"开始扫描 {len(self.stock_list)} 只股票...")
        start_time = time.time()
        
        # 使用多线程加速处理
        results = []
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(self.analyze_single_stock, stock) for stock in self.stock_list]
            
            for i, future in enumerate(futures):
                try:
                    result = future.result()
                    results.append(result)
                    
                    # 打印进度
                    if (i + 1) % 10 == 0 or (i + 1) == len(self.stock_list):
                        print(f"已处理 {i + 1}/{len(self.stock_list)} 只股票")
                except Exception as e:
                    print(f"处理股票 {self.stock_list[i]} 时出错: {e}")
        
        end_time = time.time()
        print(f"扫描完成，耗时 {end_time - start_time:.2f} 秒")
        
        # 按照是否有买入信号排序
        results.sort(key=lambda x: (not x["任一买入信号"], not x["强势突破"], not x["低吸回调"], not x["三线合一上涨"]))
        
        self.results = results
        return results
    
    def save_results(self, filename=None):
        """保存结果到Excel文件"""
        if not self.results:
            print("没有结果可保存")
            return
            
        if filename is None:
            # 使用当前日期作为文件名
            today = datetime.datetime.now().strftime("%Y%m%d")
            filename = f"股票买入信号_{today}"
            if self.new_account_only:
                filename += "_新开户可买"
            if not self.include_today:
                filename += "_不含今日数据"
            filename += ".xlsx"
        
        # 提取需要保存的字段
        data = []
        for result in self.results:
            data.append({
                "股票代码": result["股票代码"],
                "股票名称": result["股票名称"],
                "综合评分": result["综合评分"],
                "任一买入信号": result["任一买入信号"],
                "强势突破": result["强势突破"],
                "低吸回调": result["低吸回调"],
                "三线合一上涨": result["三线合一上涨"],
                "KDJ金叉": result["KDJ金叉"],
                "BIAS回归": result["BIAS回归"],
                "CCI超买超卖": result["CCI超买超卖"],
                "OBV增长": result["OBV增长"],
                "量价配合": result["量价配合"],
                "最新价格": result["最新价格"],
                "RSI": result.get("技术指标", {}).get("RSI", None),
                "KDJ_J": result.get("技术指标", {}).get("KDJ_J", None),
                "BIAS": result.get("技术指标", {}).get("BIAS", None),
                "CCI": result.get("技术指标", {}).get("CCI", None),
                "ATR": result.get("技术指标", {}).get("ATR", None),
                "SMA5": result.get("技术指标", {}).get("SMA5", None),
                "SMA10": result.get("技术指标", {}).get("SMA10", None),
                "SMA20": result.get("技术指标", {}).get("SMA20", None),
                "MACD": result.get("技术指标", {}).get("MACD", None)
            })
        
        # 创建DataFrame并处理空值
        df = pd.DataFrame(data)
        df = df.fillna("")  # 将NaN值替换为空字符串
        
        # 使用ExcelWriter保存为Excel文件
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 保存主数据表
            df.to_excel(writer, sheet_name='股票信号', index=False)
            
            # 获取工作簿和工作表对象
            workbook = writer.book
            worksheet = writer.sheets['股票信号']
            
            # 设置列宽
            for idx, col in enumerate(df.columns):
                column_width = max(len(str(col)), 12)
                worksheet.column_dimensions[chr(65 + idx)].width = column_width
            
            # 添加表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(fgColor="CCCCCC", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 为有买入信号的行添加背景色
            green_fill = PatternFill(fgColor="D8F0D8", fill_type="solid")  # 浅绿色
            
            for row_idx, row in enumerate(df.itertuples(), start=2):
                if row.任一买入信号:
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = green_fill
            
            # 为布尔值单元格设置居中对齐
            center_alignment = Alignment(horizontal='center')
            for row_idx in range(2, len(df) + 2):
                for col_idx, col_name in enumerate(df.columns, start=1):
                    if col_name in ["任一买入信号", "强势突破", "低吸回调", "三线合一上涨", "KDJ金叉", "BIAS回归", "CCI超买超卖", "OBV增长", "量价配合"]:
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = center_alignment
                        if cell.value == True:
                            cell.value = "是"
                        elif cell.value == False:
                            cell.value = "否"
            
            # 为数值单元格设置数值格式
            for row_idx in range(2, len(df) + 2):
                # 设置价格格式
                price_cell = worksheet.cell(row=row_idx, column=df.columns.get_loc("最新价格") + 1)
                if price_cell.value:
                    price_cell.number_format = "#,##0.00"
                
                # 设置评分格式
                score_cell = worksheet.cell(row=row_idx, column=df.columns.get_loc("综合评分") + 1)
                if score_cell.value:
                    score_cell.number_format = "#,##0.0"
                
                # 设置技术指标格式
                for indicator in ["RSI", "KDJ_J", "BIAS", "CCI", "MACD", "ATR", "SMA5", "SMA10", "SMA20"]:
                    if indicator in df.columns:
                        cell = worksheet.cell(row=row_idx, column=df.columns.get_loc(indicator) + 1)
                        if cell.value:
                            cell.number_format = "#,##0.00"
            
            # 添加条件格式（得分高低突出显示）
            score_col = df.columns.get_loc("综合评分") + 1
            for row_idx in range(2, len(df) + 2):
                score_cell = worksheet.cell(row=row_idx, column=score_col)
                if score_cell.value:
                    score_value = float(score_cell.value)
                    if score_value >= 50:
                        score_cell.font = Font(color="006100")  # 深绿色
                    elif score_value >= 30:
                        score_cell.font = Font(color="9C5700")  # 橙色
                    else:
                        score_cell.font = Font(color="000000")  # 黑色
            
            # 冻结首行
            worksheet.freeze_panes = "A2"
            
            # 创建信号详情工作表
            signals_df = pd.DataFrame({
                "信号类型": ["强势突破", "低吸回调", "三线合一上涨", "KDJ金叉", "BIAS回归", "CCI超买超卖", "OBV增长", "量价配合"],
                "信号权重": [
                    self.results[0].get("分析器", {}).get("权重", {}).get("强势突破", 30) if len(self.results) > 0 else 30,
                    self.results[0].get("分析器", {}).get("权重", {}).get("低吸回调", 25) if len(self.results) > 0 else 25,
                    self.results[0].get("分析器", {}).get("权重", {}).get("三线合一上涨", 20) if len(self.results) > 0 else 20,
                    self.results[0].get("分析器", {}).get("权重", {}).get("KDJ金叉", 5) if len(self.results) > 0 else 5,
                    self.results[0].get("分析器", {}).get("权重", {}).get("BIAS回归", 5) if len(self.results) > 0 else 5,
                    self.results[0].get("分析器", {}).get("权重", {}).get("CCI超买超卖", 5) if len(self.results) > 0 else 5,
                    self.results[0].get("分析器", {}).get("权重", {}).get("OBV增长", 5) if len(self.results) > 0 else 5,
                    self.results[0].get("分析器", {}).get("权重", {}).get("量价配合", 5) if len(self.results) > 0 else 5
                ],
                "信号说明": [
                    "股票呈现强势上涨趋势，均线多头排列且向上，价格站上均线，成交量配合",
                    "股价在超卖区间出现反弹迹象，RSI超卖或价格接近布林带下轨",
                    "5日、10日、20日均线高度聚合且同时向上，价格站上所有均线，预示强势上涨趋势形成",
                    "KDJ指标形成金叉或从超卖区域反弹",
                    "BIAS乖离率回归零轴或从超卖区域反弹",
                    "CCI指标从超卖区域反弹或回归零轴",
                    "OBV能量潮指标上升且价格稳定",
                    "量价配合良好，成交量增加的同时价格上涨"
                ]
            })
            
            signals_df.to_excel(writer, sheet_name='信号说明', index=False)
            
            # 美化信号说明工作表
            signal_worksheet = writer.sheets['信号说明']
            
            # 设置列宽
            for idx, col in enumerate(['A', 'B', 'C']):
                if idx == 0:
                    signal_worksheet.column_dimensions[col].width = 15
                elif idx == 1:
                    signal_worksheet.column_dimensions[col].width = 10
                else:
                    signal_worksheet.column_dimensions[col].width = 60
            
            # 添加表头样式
            for cell in signal_worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        print(f"结果已保存到 {filename}")
        return filename
    
    def print_buy_signals(self):
        """打印有买入信号的股票"""
        if not self.results:
            print("没有分析结果")
            return
            
        buy_signals = [r for r in self.results if r["任一买入信号"]]
        
        if not buy_signals:
            print("没有股票触发买入信号")
            return
            
        # 按照综合评分排序
        buy_signals.sort(key=lambda x: x.get("综合评分", 0), reverse=True)
            
        print(f"\n共有 {len(buy_signals)} 只股票触发买入信号:")
        print("-" * 120)
        
        # 构建表头
        headers = ["股票代码", "股票名称", "评分", "最新价格", "强势突破", "低吸回调", "三线合一上涨", "KDJ金叉", "BIAS回归", "CCI信号", "OBV增长", "量价配合"]
        header_format = "{:<10}{:<15}{:<8}{:<10}{:<10}{:<10}{:<10}{:<10}{:<10}{:<10}{:<10}{:<10}"
        print(header_format.format(*headers))
        
        print("-" * 120)
        
        # 打印每只股票
        row_format = "{:<10}{:<15}{:<8.1f}{:<10.2f}{:<10}{:<10}{:<10}{:<10}{:<10}{:<10}{:<10}"
        for signal in buy_signals:
            print(row_format.format(
                signal['股票代码'],
                signal['股票名称'],
                signal.get('综合评分', 0),
                signal['最新价格'],
                "✓" if signal['强势突破'] else "",
                "✓" if signal['低吸回调'] else "",
                "✓" if signal['三线合一上涨'] else "",
                "✓" if signal.get('KDJ金叉', False) else "",
                "✓" if signal.get('BIAS回归', False) else "",
                "✓" if signal.get('CCI超买超卖', False) else "",
                "✓" if signal.get('OBV增长', False) else "",
                "✓" if signal.get('量价配合', False) else ""))
        
        print("-" * 120)


def main():
    """主函数"""
    print("=" * 60)
    print("股票买入信号扫描器")
    print("=" * 60)
    
    # 选择扫描模式
    print("\n请选择扫描模式:")
    print("1. 扫描沪深300成分股")
    print("2. 扫描自定义股票列表")
    print("3. 扫描单个股票")
    print("4. 扫描所有新开户可买股票（00、000、002、60开头）")
    
    choice = input("请输入选择 (默认1): ").strip() or "1"
    
    # 询问是否包含今天的数据
    include_today = input("\n是否包含今天的数据进行计算？(y/n，默认y): ").strip().lower() != 'n'
    
    if choice == "1":
        # 扫描沪深300成分股
        scanner = StockScanner(use_index=True, index_code="000300", top_n=300, include_today=include_today)
        scanner.scan_stocks()
        scanner.print_buy_signals()
        scanner.save_results()
    
    elif choice == "2":
        # 扫描自定义股票列表
        stock_file = input("请输入股票列表文件路径 (CSV格式，包含代码和名称列): ").strip()
        
        if not os.path.exists(stock_file):
            print(f"文件 {stock_file} 不存在")
            return
            
        try:
            # 读取CSV文件
            df = pd.read_csv(stock_file)
            # 获取股票代码和名称
            stock_list = df.iloc[:, :2].values.tolist()
            
            scanner = StockScanner(stock_list=stock_list, include_today=include_today)
            scanner.scan_stocks()
            scanner.print_buy_signals()
            scanner.save_results()
        except Exception as e:
            print(f"读取股票列表文件失败: {e}")
    
    elif choice == "3":
        # 扫描单个股票
        stock_code = input("请输入股票代码: ").strip()
        stock_name = input("请输入股票名称 (可选): ").strip() or stock_code
        analyzer = BuySignalAnalyzer(stock_code, stock_name, include_today=include_today)
        
        # 使用实际数据进行完整分析
        analyzer.run_analysis()
            
        result = analyzer.get_result()
        
        print("\n分析结果:")
        print(f"股票代码: {result['股票代码']}")
        print(f"股票名称: {result['股票名称']}")
        if not include_today:
            print("注意: 分析不包含今天的数据")
        
        if result['最新价格'] is not None:
            print(f"最新价格: {result['最新价格']:.2f}")
        else:
            print("最新价格: 无数据")
            
        print(f"综合评分: {result['综合评分']:.1f}")
        print(f"任一买入信号: {'是' if result['任一买入信号'] else '否'}")
        
        # 打印详细信号
        print("\n信号触发情况:")
        print(f"强势突破: {'是' if result['强势突破'] else '否'}")
        print(f"低吸回调: {'是' if result['低吸回调'] else '否'}")
        print(f"三线合一上涨: {'是' if result['三线合一上涨'] else '否'}")
        print(f"KDJ金叉: {'是' if result['KDJ金叉'] else '否'}")
        print(f"BIAS回归: {'是' if result['BIAS回归'] else '否'}")
        print(f"CCI超买超卖: {'是' if result['CCI超买超卖'] else '否'}")
        print(f"OBV能量潮: {'是' if result['OBV增长'] else '否'}")
        print(f"量价配合: {'是' if result['量价配合'] else '否'}")
        
        # 打印技术指标值
        if "技术指标" in result and result["技术指标"] and any(v is not None for v in result["技术指标"].values()):
            print("\n主要技术指标:")
            indicators = result["技术指标"]
            for name, value in indicators.items():
                if value is not None:
                    if name == "OBV":
                        print(f"{name}: {value:.0f}")
                    else:
                        print(f"{name}: {value:.2f}")
        
        # 打印信号详情
        if result["任一买入信号"]:
            print("\n详细信号指标:")
            for signal_type, details in result["信号详情"].items():
                if result[signal_type]:
                    print(f"\n{signal_type}信号:")
                    for k, v in details.items():
                        print(f"  {k}: {v}")
                        
        # 询问是否保存为Excel
        save_excel = input("\n是否保存分析结果到Excel文件? (y/n): ").strip().lower()
        if save_excel == 'y':
            # 创建单只股票的结果列表
            single_result = [result]
            scanner = StockScanner()
            scanner.results = single_result
            filename = f"{result['股票代码']}_{result['股票名称']}_分析结果.xlsx"
            scanner.save_results(filename)
    
    elif choice == "4":
        # 扫描所有新开户可买股票
        try:
            limit = input("请输入要分析的股票数量上限（默认为300只）: ").strip()
            limit = int(limit) if limit else 300
            
            scanner = StockScanner(new_account_only=True, include_today=include_today)
            if scanner.load_new_account_stocks(limit):
                print(f"成功加载 {len(scanner.stock_list)} 只新开户可买股票")
                scanner.scan_stocks()
                scanner.print_buy_signals()
                scanner.save_results()
            else:
                print("加载新开户可买股票失败")
        except Exception as e:
            print(f"扫描新开户可买股票时出错: {e}")
            import traceback
            traceback.print_exc()
    
    else:
        print("无效的选择")


def generate_test_data():
    """生成测试数据"""
    print("生成测试数据...")
    # 生成120天的测试数据
    date_range = pd.date_range(end=pd.Timestamp.now(), periods=120)
    
    # 设置初始价格
    initial_price = 10.0
    
    # 生成随机价格序列
    np.random.seed(42)  # 设置随机种子以便结果可重现
    
    # 创建价格模拟
    daily_returns = np.random.normal(0.001, 0.02, len(date_range))
    prices = initial_price * (1 + np.cumsum(daily_returns))
    
    # 上升趋势
    prices = prices * np.linspace(1.0, 1.5, len(date_range))
    
    # 确保所有价格为正
    prices = np.maximum(prices, 0.1)
    
    # 创建高低开收价格
    high = prices * (1 + np.random.uniform(0, 0.03, len(date_range)))
    low = prices * (1 - np.random.uniform(0, 0.03, len(date_range)))
    open_prices = low + np.random.uniform(0, 1, len(date_range)) * (high - low)
    close_prices = low + np.random.uniform(0, 1, len(date_range)) * (high - low)
    
    # 创建成交量
    volume = np.random.uniform(100000, 1000000, len(date_range))
    
    # 创建数据框
    df = pd.DataFrame({
        'open': open_prices,
        'high': high,
        'low': low,
        'close': close_prices,
        'volume': volume
    }, index=date_range)
    
    return df


if __name__ == "__main__":
    main() 